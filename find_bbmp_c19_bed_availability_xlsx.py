import os
import pandas as pd
from pathlib import Path
import numpy as np
import openpyxl
import argparse
from tabulate import tabulate

parser = argparse.ArgumentParser(description='Script to generate labels file')
parser.add_argument('--bed_type', type=str, default='HDU', help='what bed type to search for; Gen, HDU, ICU, ICU Ventl')

args = parser.parse_args()


def check_merged_cell(sheet_obj, cell_xy):
    
    cell = sheet_obj.cell(cell_xy[1], cell_xy[0])
    for mergedCell in sheet_obj.merged_cells.ranges:
        if cell.coordinate in mergedCell:
            return True
    return False

def find_merged_cell_col_range(sheet_obj, total_rows_cols, cell_xy):

    col_range = []
    
    # Store a list of all the merged columns
    merged_cells_bb_tlbr = []
    for mergedCell in sheet_obj.merged_cells.ranges:
        
        min_x = mergedCell.min_col
        max_x = mergedCell.max_col
        min_y = mergedCell.min_row
        max_y = mergedCell.max_row
        
        m_cell_xy = np.array([min_x, min_y])
        
        if np.array_equal(m_cell_xy, cell_xy):
            col_range = [min_x, max_x]
            
    return col_range

def find_req_tables_start_xy(sheet_obj, total_rows_cols, req_table_col_header):
    
    req_tables_start_xy = []
    for r_idx in range(1, total_rows_cols[0]+1):
        for c_idx in range(1, total_rows_cols[1]+1):
            cell_val = sheet_obj.cell(row=r_idx, column=c_idx).value
            if cell_val is None:
                continue
            if cell_val == req_table_col_header:
                req_tables_start_xy.append([c_idx,r_idx])
            
    return req_tables_start_xy

def find_hospital_type(sheet_obj, req_tables_start_xy):
    
    hospital_types = []
    
    for t_xy in req_tables_start_xy:
        cell_val = sheet_obj.cell(row=t_xy[1]-2, column=1).value
        if cell_val is None:
            continue
        hospital_types.append(cell_val)
        
    return hospital_types

def find_bed_type_col_idx(sheet_obj, row_range, col_range, bed_type):
    
    req_col_idx = -1
    
    bed_type_splits = bed_type.split(' ')
    bed_type_len = len(bed_type_splits)
    
    for r_idx in range(row_range[0], row_range[1]+1):
        for c_idx in range(col_range[0], col_range[1]+1):
            
            bed_type_matches = 0
            for b_idx, bed_type in enumerate(bed_type_splits):
                cell_val = sheet_obj.cell(row=r_idx+b_idx, column=c_idx).value
                if cell_val is None:
                    bed_type_matches = 0
                    continue
                if cell_val == bed_type:
                    bed_type_matches += 1
                else:
                    bed_type_matches = 0
                    
            if bed_type_matches == bed_type_len:
                req_col_idx = c_idx
                break

        if req_col_idx >= 0:
            break
        
    return req_col_idx
    
def find_bed_availability(sheet_obj, row_range, col_idx):
    
    valid_row_idxs = []
    
    for r_idx in range(row_range[0], row_range[1]):
        cell_val = sheet_obj.cell(row=r_idx, column=col_idx).value
        if cell_val is None:
            continue
        if cell_val > 0:
            valid_row_idxs.append(r_idx)
            
    return valid_row_idxs
    
def display_bed_availability(sheet_obj, valid_row_idxs, disp_col_idxs, hospital_type):
    
    table_vals = []
    for r_idx in valid_row_idxs:
        
        row_vals = []
        for c_idx in disp_col_idxs:
            cell = sheet_obj.cell(row=r_idx, column=c_idx)
            cell_val = cell.value
            if cell_val is None:
                continue
            row_vals.append(cell_val)
            
        table_vals.append(row_vals)
    
    if len(table_vals):
        print('')
        print('Hospital Type: %s\n' % (hospital_type))
        print(tabulate(table_vals, headers=["SlNo", "Hospital", "Availability"], numalign="center", stralign="center"))
        print('')
        
f_path = Path.cwd() # change to the path to the Excel file
f_name = 'bbmp_covid19_bed_status.xlsx' # Excel File name
filename = os.path.join(f_path, f_name)
sheetname = '20210514' # change to the name of the worksheet

# To open the workbook 
# workbook object is created
wb_obj = openpyxl.load_workbook(filename)

# Get workbook active sheet object
# from the active attribute
sheet_obj = wb_obj[sheetname]

total_rows_cols = [sheet_obj.max_row, sheet_obj.max_column]

start_row = 1
last_row = total_rows_cols[0]

bed_type = args.bed_type

req_table_col_header = 'Net Available Beds for C+ Patients'
req_tables_start_xy = find_req_tables_start_xy(sheet_obj, total_rows_cols, req_table_col_header)

hospital_types = find_hospital_type(sheet_obj, req_tables_start_xy)

req_tables_start_xy.append([0, total_rows_cols[0]+1])
for t_idx,t_xy in enumerate(req_tables_start_xy[:-1]):
    
    row_range = [t_xy[1]+1, req_tables_start_xy[t_idx+1][1]]

    if check_merged_cell(sheet_obj, t_xy):
        col_range = find_merged_cell_col_range(sheet_obj, total_rows_cols, t_xy)
    else:
        col_range = [t_xy[0], t_xy[0]]

    req_col_idx = find_bed_type_col_idx(sheet_obj, row_range, col_range, bed_type)
    if req_col_idx == -1:
        continue
    
    row_range[0] += 2
    valid_row_idxs = find_bed_availability(sheet_obj, row_range, req_col_idx)
    if len(valid_row_idxs) == 0:
        continue

    disp_col_idxs = [1, 2, req_col_idx]
    display_bed_availability(sheet_obj, valid_row_idxs, disp_col_idxs, hospital_types[t_idx])
